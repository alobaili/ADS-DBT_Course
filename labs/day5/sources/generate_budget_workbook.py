"""
Build the district budget workbook used by the Day 5 ingestion lab.

This is deliberately shaped like a spreadsheet a finance team would actually
send you, rather than a tidy machine-readable export:

  * one sheet per quarter, so the reader has to loop over sheets
  * three rows of title and provenance text above the real header row
  * a blank spacer column between the two column groups
  * a total row at the foot of each sheet that must not be loaded as data
  * budget written as text with a currency symbol and thousands separators
  * district names rather than district ids, so the join needs the dimension

Every one of those is a real defect that appears in real workbooks, and each is
handled explicitly in the lab.

Run:
    python generate_budget_workbook.py
Writes:
    district_budgets.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

DISTRICTS = [
    ("Al Olaya", 210000),
    ("Al Malaz", 185000),
    ("Al Naseem", 240000),
    ("Irqah", 96000),
    ("Al Aziziyah", 158000),
    ("Diriyah", 74000),
]

# Fixed figures so every delegate computes identical numbers. Budget in
# thousands, headcount as whole people.
QUARTERS = {
    "2024-Q1": [(412, 18), (368, 16), (455, 21), (198, 9), (301, 14), (162, 7)],
    "2024-Q2": [(428, 19), (372, 16), (471, 22), (203, 9), (315, 14), (168, 7)],
    "2024-Q3": [(441, 19), (359, 15), (488, 23), (211, 10), (322, 15), (171, 8)],
    "2024-Q4": [(465, 20), (381, 17), (502, 24), (219, 10), (338, 15), (177, 8)],
    "2025-Q1": [(478, 21), (394, 17), (516, 24), (226, 11), (349, 16), (183, 8)],
}


def build_sheet(ws, quarter, rows):
    bold = Font(bold=True)

    # Three rows of preamble before the real header, which is row 4.
    ws["A1"] = "Municipal Finance Office"
    ws["A1"].font = bold
    ws["A2"] = "District Operating Budget and Establishment"
    ws["A3"] = "Period: %s   Status: Final   Prepared by: Finance Systems" % quarter

    # Real header row. Column C is a deliberate blank spacer.
    ws["A4"] = "District"
    ws["B4"] = "Operating Budget"
    ws["C4"] = None
    ws["D4"] = "Headcount"
    for cell in ("A4", "B4", "D4"):
        ws[cell].font = bold

    r = 5
    total_budget = 0
    total_head = 0
    for (name, _pop), (budget, head) in zip(DISTRICTS, rows):
        ws.cell(row=r, column=1, value=name)
        # Written as text with a symbol and separators, exactly as finance send it.
        ws.cell(row=r, column=2, value="SAR {:,}".format(budget * 1000))
        ws.cell(row=r, column=4, value=head)
        total_budget += budget * 1000
        total_head += head
        r += 1

    # Total row, which must be excluded on load.
    ws.cell(row=r, column=1, value="TOTAL").font = bold
    ws.cell(row=r, column=2, value="SAR {:,}".format(total_budget)).font = bold
    ws.cell(row=r, column=4, value=total_head).font = bold

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 14


def main():
    wb = Workbook()
    wb.remove(wb.active)
    for quarter, rows in QUARTERS.items():
        ws = wb.create_sheet(title=quarter)
        build_sheet(ws, quarter, rows)

    # A non-data sheet, so "read every sheet" is not a safe rule.
    notes = wb.create_sheet(title="Notes")
    notes["A1"] = "Figures are indicative and subject to in-year revision."
    notes["A2"] = "Budget shown in Saudi Riyal. Headcount is funded establishment."

    out = Path(__file__).resolve().parent / "district_budgets.xlsx"
    wb.save(out)
    print("wrote %s with sheets: %s" % (out, ", ".join(wb.sheetnames)))


if __name__ == "__main__":
    main()
